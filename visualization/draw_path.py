import os
import json
import numpy as np
import open3d as o3d
from math import sqrt, pi, cos, sin
from openpyxl import load_workbook
from typing import List


case_name = 'sca'

abs_path = os.path.abspath('.')
log_save_dir = abs_path + '/' + case_name + '/'


def load_path_data(filename, agents_num) -> dict:
    agent_trajs = dict()
    wb = load_workbook(filename)
    for index in range(agents_num):
        exec('agent' + str(index) + '= read_sheet(wb, \'agent' + str(index) + '\')')
        exec('agent_trajs[\'agent' + str(index) + '\'] = dict2array_pos(agent' + str(index) + ')')
    return agent_trajs


def read_sheet(wb, sheet_name) -> List[dict]:
    sheet = wb[sheet_name]
    row_len = sheet.max_row
    column_len = sheet.max_column
    data = []
    for row in range(2, row_len + 1):
        sub_data = dict()
        sub_data['index'] = int(row) - 2
        for column in range(2, column_len + 1):
            sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value
        data.append(sub_data)
    return data


def dict2array_pos(data: list) -> np.array:
    data_list = []
    for i in range(len(data)):
        data_list.append([data[i]['pos_x'], data[i]['pos_y'], data[i]['pos_z']])
    return np.array(data_list)


def points_obs(obstacles: list) -> np.array:
    obs_list = list()
    for obs in obstacles:
        obs_list.append(obs.pos_global_frame)
    return np.array(obs_list)


def draw_sphere(center, rad=1.0):
    num_N = int(rad*512)
    obj_pos = []
    param_phi = (sqrt(5.0) - 1.0) / 2.0
    for i in range(1, num_N + 1):
        z_n = (2 * i - 1) / num_N - 1
        x_n = sqrt(1 - z_n ** 2) * cos(2 * pi * i * param_phi)
        y_n = sqrt(1 - z_n ** 2) * sin(2 * pi * i * param_phi)
        pos = [rad * x_n + center[0], rad * y_n + center[1], rad * z_n + center[2]]
        obj_pos.append(pos)

    return obj_pos


def read_obstacles_pos(obstacles_information):
    obs_list = list()
    for i in range(len(obstacles_information)):
        feature = obstacles_information[i]['feature']
        position = obstacles_information[i]['position']
        if feature <= 0.2:
            obs_list.append(position)
        else:
            obs_list += draw_sphere(position, rad=feature)
    return np.array(obs_list)


def get_agent_info(filename):
    with open(filename, "r") as f:
        json_info = json.load(f)

    agents_information = json_info['all_agent_info']
    obstacles_information = json_info['all_obstacle']

    return agents_information,  obstacles_information


if __name__ == '__main__':
    info_file = log_save_dir + 'log/env_cfg.json'
    traj_file = log_save_dir + 'log/trajs.xlsx'

    agents_info, obstacles_info = get_agent_info(info_file)
    agent_num = len(agents_info)
    uav_trajs = load_path_data(traj_file, agent_num)

    pcd_list = list()
    for n in range(len(uav_trajs)):
        pcd = o3d.geometry.PointCloud()
        pcd.points = o3d.utility.Vector3dVector(uav_trajs['agent' + str(n)])
        pcd_list.append(pcd)

    obstacles_pos = read_obstacles_pos(obstacles_info)
    pcd = o3d.geometry.PointCloud()
    pcd.points = o3d.utility.Vector3dVector(obstacles_pos)
    pcd_list.append(pcd)

    o3d.visualization.draw_geometries([pcd for pcd in pcd_list])
